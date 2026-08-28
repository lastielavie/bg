"""
Dashboard Bagi Hasil Teknisi
============================
Jalankan:
    streamlit run app.py
"""
import io
import re
from copy import copy
from datetime import date
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
import streamlit as st

st.set_page_config(page_title="Bagi Hasil Teknisi", layout="wide", page_icon="🧰")

TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"

# Deteksi file master final (toleran huruf besar / kecil: final.xlsx, FINAL.xlsx)
FINAL_PATH = None
for _p in [Path(__file__).parent / "FINAL.xlsx", Path(__file__).parent / "final.xlsx"]:
    if _p.exists():
        FINAL_PATH = _p
        break
if FINAL_PATH is None:
    FINAL_PATH = Path(__file__).parent / "FINAL.xlsx"

BULAN_NAMES = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli',
               'Agustus', 'September', 'Oktober', 'November', 'Desember']

PALETTE = ['#1f3864', '#2e9bd6', '#16a34a', '#e0921f', '#c9392f',
           '#7c3aed', '#0f8a82', '#a855f7', '#3f8ac9', '#d1478d']

KATA_KUNCI_TARIF = ['INTERFACE', 'NORMAL', 'MATI TOTAL', 'PROMO']
KATEGORI_TARIF = ['Interface', 'Normal', 'Mati Total', 'Promo', 'Lainnya']
TARIF_AWAL = {'Interface': 20.0, 'Normal': 30.0, 'Mati Total': 32.0, 'Promo': 60.0}
TARIF_DEFAULT_AWAL = 30.0
TARIF_PEMBANDING_AWAL = 30.0
LABEL_LAINNYA = 'Lainnya'

POLA_JASA_DIKECUALIKAN = ['OPER GADGET']

CABANG_ACUAN_KERUSAKAN = ['CONDET']
KERUSAKAN_INTERFACE = ['BATERAI', 'SSD', 'RAM']
KERUSAKAN_NORMAL = ['FLEXIBEL', 'FLEXIBLE', 'FLEKSIBEL', 'MIC', 'WIFI CARD', 'REPAIR']


def norm_formula(val):
    """Menyesuaikan rumus Excel agar kompatibel dengan openpyxl & Excel XML."""
    if isinstance(val, str) and val.startswith('='):
        # 1. Hapus prefiks _xludf. jika otomatis ditambahkan oleh Excel
        val = re.sub(r'_xludf\.', '', val, flags=re.I)

        # 2. Pastikan fungsi Excel modern memiliki prefix _xlfn.
        modern_funcs = ['TEXTJOIN', 'IFS', 'CONCAT', 'XLOOKUP', 'XMATCH', 'SWITCH', 'MAXIFS', 'MINIFS']
        for fn in modern_funcs:
            val = re.sub(rf'(?<!_xlfn\.)\b{fn}\b', f'_xlfn.{fn}', val, flags=re.I)
        
        # 3. Ganti titik koma (;) menjadi koma (,) di luar string (tanda petik)
        parts = re.split(r'("[^"]*")', val)
        for i in range(0, len(parts), 2):
            parts[i] = parts[i].replace(';', ',')
        val = "".join(parts)
    return val


def label_dari_kerusakan(kerusakan, kategori_jual):
    ku = str(kerusakan or '').upper()
    kp = str(kategori_jual or '').upper()
    laptop = 'LAPTOP' in kp
    if 'MATI TOTAL' in ku:
        return 'Mati Total'
    if 'LCD' in ku:
        return 'Normal' if laptop else ('Interface' if 'HP' in kp else 'Normal')
    if 'SOFTWARE' in ku:
        return 'Interface' if laptop else 'Normal'
    if any(k in ku for k in KERUSAKAN_INTERFACE):
        return 'Interface'
    return 'Normal'


NAMA_TARIF_TETAP_20 = [
    'M IBNU SIDIK', 'RAFI ALAMSYAH', 'MIFTAHUL MUTTAQIEN', 'BRYAN PUTRA',
    'HAMZAH MAULANA', 'DAVID SONDAKH', 'FATHUR ROHMAN SOBARNA', 'ALAI ARKAN',
    'ALFIN DAMARA', 'ADI FIRDAUS', 'M IQBAL PRADANA', 'IRSYAD PANCA GUNAWAN',
    'EGI SETYA RAMADHANI', 'ALVI SYAHLANI RAMADHAN', 'M NAUFAL', 'M HANIF FATIN',
    'SYAHDAN IBNU FAUZI', 'JUNARA', 'KARIM AGAKHAN', 'FARID HASBY ASH SHIDDIQ',
    'RIDWAN KURNIAWAN',
]
NAMA_TARIF_NORMAL_355 = ['EKO RISDIYANTORO JATIMULYA', 'IRVAN SYAHRONI']
KOL_TARIF_KHUSUS = ['Nama Teknisi', 'Interface', 'Normal', 'Mati Total', 'Promo',
                    'Lainnya']


def tarif_khusus_awal():
    baris = [{'Nama Teknisi': n, 'Interface': 20.0, 'Normal': 20.0,
              'Mati Total': 22.0, 'Promo': None, 'Lainnya': 20.0}
             for n in NAMA_TARIF_TETAP_20]
    baris += [{'Nama Teknisi': n, 'Interface': None, 'Normal': 35.5,
               'Mati Total': 37.5, 'Promo': None, 'Lainnya': None}
              for n in NAMA_TARIF_NORMAL_355]
    return pd.DataFrame(baris, columns=KOL_TARIF_KHUSUS)


def _nama_rapi(s):
    return re.sub(r'\s+', ' ', str(s).strip().upper())


def peta_tarif_khusus(tabel):
    hasil = {}
    if tabel is None or len(tabel) == 0:
        return hasil
    for _, r in tabel.iterrows():
        nama = _nama_rapi(r.get('Nama Teknisi', ''))
        if not nama or nama == 'NAN':
            continue
        tar = {}
        for lbl in KATEGORI_TARIF:
            v = r.get(lbl)
            if v is not None and not pd.isna(v):
                tar[lbl] = float(v) / 100.0
        if tar:
            hasil[nama] = tar
    return hasil


def cocokkan_teknisi(nama_teknisi, kunci_khusus):
    peta = {}
    kunci = sorted(kunci_khusus, key=len, reverse=True)
    for t in nama_teknisi:
        tn = _nama_rapi(t)
        for k in kunci:
            if tn == k or tn.startswith(k + ' '):
                peta[t] = k
                break
    return peta

st.markdown("""
<style>
  .kpi-wrap{display:grid;grid-template-columns:repeat(6,1fr);gap:14px;margin-bottom:6px;}
  @media(max-width:1200px){.kpi-wrap{grid-template-columns:repeat(3,1fr);}}
  .kpi{border-radius:16px;padding:16px 16px 18px;color:#fff;min-height:112px;
       box-shadow:0 8px 20px rgba(30,20,60,.14);}
  .kpi .label{font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.04em;opacity:.92;}
  .kpi .value{font-size:21px;font-weight:800;margin-top:10px;line-height:1.15;}
  .kpi .foot{font-size:11px;margin-top:6px;opacity:.9;}
</style>
""", unsafe_allow_html=True)


def rp(v, singkat=True):
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "-"
    neg, v = v < 0, abs(float(v))
    if singkat:
        if v >= 1_000_000_000:
            s = f"Rp {v/1_000_000_000:,.2f} M"
        elif v >= 1_000_000:
            s = f"Rp {v/1_000_000:,.1f} jt"
        elif v >= 1_000:
            s = f"Rp {v/1_000:,.0f} rb"
        else:
            s = f"Rp {v:,.0f}"
    else:
        s = f"Rp {v:,.0f}"
    s = s.replace(",", "#").replace(".", ",").replace("#", ".")
    return ("-" + s) if neg else s


def kpi_html(cards):
    cells = []
    for c in cards:
        cells.append(f"""
        <div class="kpi" style="background:{c['grad']}">
          <div class="label">{c['label']}</div>
          <div class="value">{c['value']}</div>
          <div class="foot">{c.get('sub','&nbsp;')}</div>
        </div>""")
    return f'<div class="kpi-wrap">{"".join(cells)}</div>'


def cocok_kata_kunci(nama_barang):
    s = str(nama_barang).upper()
    return [k for k in KATA_KUNCI_TARIF if k in s]


def pilih_label_tarif(kw_str, urutan):
    if not kw_str:
        return LABEL_LAINNYA
    cocok = kw_str.split('|')
    for k in urutan:
        if k in cocok:
            return k.title()
    return cocok[0].title()


def periode_gaji(bulan_gaji: int, tahun_gaji: int):
    m_akhir, th_akhir = bulan_gaji, tahun_gaji
    m_awal, th_awal = m_akhir - 1, th_akhir
    if m_awal < 1:
        m_awal += 12
        th_awal -= 1
    return pd.Timestamp(th_awal, m_awal, 24), pd.Timestamp(th_akhir, m_akhir, 23)


def label_periode(bulan_gaji, tahun_gaji):
    a, b = periode_gaji(bulan_gaji, tahun_gaji)
    return (f"Gaji {BULAN_NAMES[bulan_gaji]} {tahun_gaji}  "
            f"({a.day} {BULAN_NAMES[a.month]} – {b.day} {BULAN_NAMES[b.month]} {b.year})")


def daftar_periode_gaji(tgl_min, tgl_max):
    hasil = []
    if pd.isna(tgl_min) or pd.isna(tgl_max):
        return hasil
    y, m = tgl_min.year, tgl_min.month
    for _ in range(120):
        a, b = periode_gaji(m, y)
        if a > tgl_max:
            break
        if b >= tgl_min:
            hasil.append((y, m))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return hasil


SALES_REQUIRED = ['TGL FAKTUR', 'NO FAKTUR', 'KATEGORI BARANG', 'NAMA BARANG',
                  'QTY', 'TOTAL HARGA']
KOLOM_DIPAKAI = SALES_REQUIRED + ['CABANG', 'NAMA TEKNISI', 'NAMA TEKNISI (FINAL)',
                                 'KERUSAKAN UTAMA', 'KATEGORI PENJUALAN']
KUNCI_DUPLIKAT = ['CABANG', 'NO FAKTUR', 'NAMA BARANG', 'QTY', 'TOTAL HARGA',
                  'TGL FAKTUR']
LABEL_TANPA_CABANG = '(TANPA CABANG)'

CABANG_KANONIK = [
    'BINTARA', 'CEGER', 'CIBINONG', 'CIBUBUR', 'CIKAMPEK', 'CILANGKAP', 'CINERE',
    'CONDET', 'DRAMAGA', 'JATIBENING', 'JATIMULYA', 'JATIWARINGIN', 'KARAWANG',
    'KLENDER', 'PEJATEN', 'RADJIMAN', 'SAWANGAN', 'WARBONG',
]
ALIAS_CABANG_AWAL = {'TELUK JAMBE': 'KARAWANG', 'TELUKJAMBE': 'KARAWANG'}

NOISE_NAMA = {
    'RINCIAN', 'PENJUALAN', 'PENJUALANAN', 'DATA', 'LAPORAN', 'LAP', 'REKAP', 'JASA',
    'SALES', 'FAKTUR', 'INVOICE', 'PERIODE', 'BULAN', 'TAHUN', 'CABANG', 'CAB',
    'BRANCH', 'TOKO', 'FIX', 'FINAL', 'REVISI', 'REV', 'UPDATE', 'BARU', 'NEW',
    'COPY', 'SALINAN', 'XLSX', 'XLS', 'CSV', 'GZ', 'FILE', 'BAGI', 'HASIL',
    'TEKNISI', 'OK', 'SHEET', 'LEMBAR', 'TABEL', 'TABLE',
    'JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI', 'JULI', 'AGUSTUS',
    'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER',
    'JAN', 'FEB', 'MAR', 'APR', 'JUN', 'JUL', 'AGU', 'AGS', 'SEP', 'OKT', 'NOV', 'DES',
}
AWALAN_BUANG = ('MFLASH', 'MFLSH', 'MFLAS')


def _huruf(s) -> str:
    return re.sub(r'[^A-Z]', '', str(s).upper())


def token_cabang(nama: str) -> str:
    s = re.sub(r'\.(xlsx|xlsm|xls|csv|gz|txt)$', '', str(nama), flags=re.I)
    s = re.sub(r'\.(csv|xlsx)$', '', s, flags=re.I)
    s = re.sub(r'\b[0-9a-f]{8,}\b', ' ', s, flags=re.I)
    s = re.sub(r'[\_\-\.\(\)\[\]#]+', ' ', s)
    s = re.sub(r'\d+', ' ', s)
    tok = []
    for t in s.upper().split():
        for aw in AWALAN_BUANG:
            if t.startswith(aw) and len(t) > len(aw):
                t = t[len(aw):]
        if len(t) > 1 and t not in NOISE_NAMA:
            tok.append(t)
    return ' '.join(tok).strip()


def cocokkan_cabang(tok: str, kanonik, alias):
    t = _huruf(tok)
    if not t:
        return '', ''
    for a, v in alias.items():
        A = _huruf(a)
        if A and (A.startswith(t) or t.startswith(A)):
            return v, f'alias {a}'
    kandidat = [c for c in kanonik
                if _huruf(c).startswith(t) or t.startswith(_huruf(c))]
    if len(kandidat) == 1:
        return kandidat[0], ''
    if len(kandidat) > 1:
        return tok.upper(), 'ambigu: ' + '/'.join(kandidat)
    return tok.upper(), 'di luar daftar'


def _cabang_dari_kolom(d: pd.DataFrame) -> str:
    if 'CABANG' not in d.columns:
        return ''
    v = d['CABANG'].dropna().astype(str).str.strip()
    v = v[(v != '') & (~v.str.upper().isin(['NAN', 'NONE']))]
    if v.empty:
        return ''
    return '' if v.nunique() > 1 else v.iloc[0]


def _potongan_berkas(nama_berkas: str, isi: bytes):
    low = str(nama_berkas).lower()
    if low.endswith('.gz'):
        yield pd.read_csv(io.BytesIO(isi), compression='gzip'), ''
    elif low.endswith(('.csv', '.txt')):
        yield pd.read_csv(io.BytesIO(isi)), ''
    else:
        xls = None
        for mesin in ('calamine', 'openpyxl'):
            try:
                xls = pd.ExcelFile(io.BytesIO(isi), engine=mesin)
                break
            except Exception:
                continue
        if xls is None:
            xls = pd.ExcelFile(io.BytesIO(isi), engine='openpyxl')
        for sheet in xls.sheet_names:
            if str(sheet).strip().lower() in ['pivot', 'pivottable', 'sheet n', 'pivot table']:
                continue
            d = xls.parse(sheet)
            if not d.empty:
                yield d, sheet


@st.cache_data(show_spinner="Membaca berkas penjualan...")
def baca_mentah(items: tuple, kanonik: tuple, alias_items: tuple):
    alias = dict(alias_items)
    frames, catatan, gagal = [], [], []
    for nama_berkas, isi in items:
        try:
            potongan = list(_potongan_berkas(nama_berkas, isi))
        except Exception as e:
            gagal.append(f"{nama_berkas}: {e}")
            continue
        if not potongan:
            gagal.append(f"{nama_berkas}: tidak ada baris data transaksi")
            continue
        for d, bagian in potongan:
            kurang = [c for c in SALES_REQUIRED if c not in d.columns]
            if kurang:
                if len(potongan) == 1:
                    gagal.append(f"{nama_berkas}"
                                 + (f" [{bagian}]" if bagian else "")
                                 + ": kolom tidak ditemukan — " + ", ".join(kurang))
                continue
            d = d[[c for c in KOLOM_DIPAKAI if c in d.columns]].copy()

            cab, asal, ket = _cabang_dari_kolom(d), 'kolom CABANG', ''
            if not cab and bagian:
                tok = token_cabang(bagian)
                if tok:
                    cab, ket = cocokkan_cabang(tok, kanonik, alias)
                    asal = 'nama sheet'
            if not cab:
                tok = token_cabang(nama_berkas)
                if tok:
                    cab, ket = cocokkan_cabang(tok, kanonik, alias)
                    asal = 'nama berkas'
            beragam = ('CABANG' in d.columns) and bool(d['CABANG'].notna().any())
            if not cab:
                asal, ket = ('kolom CABANG', 'beragam per baris') if beragam \
                    else ('—', 'perlu diisi manual')

            if cab:
                d['CABANG'] = cab
            elif not beragam:
                d['CABANG'] = pd.NA
            d['__BERKAS__'] = nama_berkas
            d['__BAGIAN__'] = bagian
            frames.append(d)
            catatan.append({
                'Berkas': nama_berkas, 'Bagian': bagian or '—',
                'Cabang': cab if cab else ('(beragam)' if beragam else LABEL_TANPA_CABANG),
                'Dari': asal, 'Catatan': ket, 'Baris': len(d)})
    if not frames:
        return pd.DataFrame(), pd.DataFrame(catatan), gagal
    return (pd.concat(frames, ignore_index=True, sort=False),
            pd.DataFrame(catatan), gagal)


NAMA_KOSONG = {'', '-', '--', 'NAN', 'NONE', 'NULL', '<NA>', 'N/A', 'NA', '#N/A',
               'N.A', 'N.A.', '#VALUE!', '#REF!', 'TIDAK ADA'}


def _nama_teknisi_bersih(kolom: pd.Series) -> pd.Series:
    s = (kolom.astype(str).str.replace(r'\s+', ' ', regex=True)
         .str.strip().str.upper().fillna(''))
    return s.where(~s.isin(NAMA_KOSONG), '')


def bersihkan(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    for c in ['QTY', 'TOTAL HARGA']:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    df['TGL'] = pd.to_datetime(df['TGL FAKTUR'], errors='coerce')
    df['KATEGORI'] = df['KATEGORI BARANG'].astype(str).str.strip().str.upper()
    df['BARANG'] = df['NAMA BARANG'].astype(str).str.strip()
    df['CABANG'] = (df['CABANG'].astype(str).str.strip()
                    .replace({'': LABEL_TANPA_CABANG, 'nan': LABEL_TANPA_CABANG,
                              'NaN': LABEL_TANPA_CABANG, 'None': LABEL_TANPA_CABANG})
                    .fillna(LABEL_TANPA_CABANG))

    fin = _nama_teknisi_bersih(df['NAMA TEKNISI (FINAL)']) if 'NAMA TEKNISI (FINAL)' in df.columns \
        else pd.Series('', index=df.index)
    asli = _nama_teknisi_bersih(df['NAMA TEKNISI']) if 'NAMA TEKNISI' in df.columns \
        else pd.Series('', index=df.index)

    df['TEKNISI'] = fin.where(fin != '', asli)
    df.loc[df['TEKNISI'] == '', 'TEKNISI'] = 'TIDAK ADA TEKNISI'

    df = df[df['KATEGORI'] == 'JASA'].copy()
    for asal, baru in [('KERUSAKAN UTAMA', 'KERUSAKAN'),
                       ('KATEGORI PENJUALAN', 'KAT_JUAL')]:
        df[baru] = (df[asal].astype(str).str.replace(r'\s+', ' ', regex=True)
                    .str.strip().str.upper().fillna('')
                    if asal in df.columns else '')
        df.loc[df[baru].isin(['NAN', 'NONE', '<NA>']), baru] = ''
    df['KW_MATCH'] = df['BARANG'].map(lambda s: '|'.join(cocok_kata_kunci(s)))
    return df


st.sidebar.title("📁 Sumber Data")
up = st.sidebar.file_uploader(
    "Upload data penjualan",
    type=['xlsx', 'xlsm', 'gz', 'csv'],
    accept_multiple_files=False,
    key='uploader_cabang',
    help="Upload 1 berkas penjualan gabungan.")

buang_duplikat = st.sidebar.checkbox(
    "Buang kiriman ulang (duplikat)", value=True, key='opsi_dedup',
    help="Baris yang sama persis (cabang, no faktur, barang, qty, total, tanggal) hanya dihitung sekali.")

with st.sidebar.expander("🏷️ Daftar & alias cabang", expanded=False):
    st.caption("Nama berkas/sheet dicocokkan ke daftar di bawah lewat awalan nama.")
    st.session_state.setdefault('teks_kanonik', "\n".join(CABANG_KANONIK))
    st.session_state.setdefault(
        'teks_alias', "\n".join(f"{k} = {v}" for k, v in ALIAS_CABANG_AWAL.items()))
    teks_kanonik = st.text_area(
        "Daftar cabang resmi (satu per baris)", height=150, key='teks_kanonik')
    teks_alias = st.text_area(
        "Alias — format `nama lain = CABANG RESMI`", height=100, key='teks_alias')

kanonik = tuple(dict.fromkeys(
    b.strip().upper() for b in teks_kanonik.splitlines() if b.strip()))
alias_items = tuple(
    (a.strip().upper(), b.strip().upper())
    for a, _, b in (baris.partition('=') for baris in teks_alias.splitlines())
    if a.strip() and b.strip())

jasa_all = pd.DataFrame()
catatan_berkas = pd.DataFrame()
raw_uploaded_bytes = None

try:
    if up is not None:
        raw_uploaded_bytes = up.getvalue()
        items = ((up.name, raw_uploaded_bytes),)
        mentah, catatan_berkas, gagal = baca_mentah(items, kanonik, alias_items)
        mentah = mentah.copy()

        if not mentah.empty and not catatan_berkas.empty:
            perlu = catatan_berkas['Cabang'] == LABEL_TANPA_CABANG
            if perlu.any():
                st.sidebar.warning(f"{int(perlu.sum())} bagian belum ketahuan cabangnya — "
                                   "isi manual di bawah.")
                with st.sidebar.form('form_cabang'):
                    isian = {}
                    for _, r in catatan_berkas[perlu].iterrows():
                        label = r['Berkas'] + (f" [{r['Bagian']}]" if r['Bagian'] != '—' else '')
                        isian[(r['Berkas'], r['Bagian'])] = st.text_input(
                            f"Cabang untuk {label}", key=f"cab_{label}")
                    if st.form_submit_button("Terapkan nama cabang"):
                        st.session_state['peta_cabang'] = {
                            k: v.strip().upper() for k, v in isian.items() if v.strip()}
                        st.rerun()
            for (berkas, bagian), cab in st.session_state.get('peta_cabang', {}).items():
                m = (mentah['__BERKAS__'] == berkas) & \
                    (mentah['__BAGIAN__'] == ('' if bagian == '—' else bagian))
                mentah.loc[m, 'CABANG'] = cab
                mc = (catatan_berkas['Berkas'] == berkas) & (catatan_berkas['Bagian'] == bagian)
                catatan_berkas.loc[mc, 'Cabang'] = cab
                catatan_berkas.loc[mc, 'Dari'] = 'isian manual'
                catatan_berkas.loc[mc, 'Catatan'] = ''

        n_dup = 0
        if not mentah.empty and buang_duplikat and \
                all(c in mentah.columns for c in KUNCI_DUPLIKAT):
            sebelum = len(mentah)
            mentah['__N__'] = mentah.groupby(KUNCI_DUPLIKAT + ['__BERKAS__'],
                                             dropna=False).cumcount()
            mentah = mentah.drop_duplicates(subset=KUNCI_DUPLIKAT + ['__N__'],
                                            keep='last').drop(columns='__N__')
            n_dup = sebelum - len(mentah)

        jasa_all = bersihkan(mentah)
        if gagal:
            st.sidebar.error("Berkas dilewati:\n\n- " + "\n- ".join(gagal))
        if not jasa_all.empty:
            st.sidebar.success(
                f"1 berkas · {jasa_all['CABANG'].nunique()} cabang · "
                f"{len(jasa_all):,} baris jasa"
                + (f" · {n_dup:,} baris duplikat dibuang" if n_dup else ""))
except Exception as e:
    st.sidebar.error(f"Data tidak terbaca: {e}")

if not catatan_berkas.empty:
    with st.sidebar.expander(f"📄 Rincian berkas terbaca", expanded=True):
        st.dataframe(catatan_berkas, hide_index=True, use_container_width=True)

st.title("🧰 Bagi Hasil Teknisi")

if jasa_all.empty:
    st.info("Data belum tersedia. Silakan upload file penjualan lewat panel kiri.")
    st.stop()

st.caption(
    f"{len(jasa_all):,} baris jasa · {jasa_all['CABANG'].nunique()} cabang · "
    f"{jasa_all.loc[jasa_all['TEKNISI'] != 'TIDAK ADA TEKNISI', 'TEKNISI'].nunique()} teknisi · "
    f"data {jasa_all['TGL'].min():%d %b %Y} – {jasa_all['TGL'].max():%d %b %Y}"
)

with st.expander("⚙️ Pengaturan Tarif Bagi Hasil — klik untuk mengubah", expanded=False):
    c1, c2, c3, c4 = st.columns(4)
    tarif_input = {}
    with c1:
        tarif_input['Interface'] = st.number_input(
            "Interface (%)", 0.0, 100.0, TARIF_AWAL['Interface'], 1.0, key='t_int')
    with c2:
        tarif_input['Normal'] = st.number_input(
            "Normal (%)", 0.0, 100.0, TARIF_AWAL['Normal'], 1.0, key='t_nor')
    with c3:
        tarif_input['Mati Total'] = st.number_input(
            "Mati Total (%)", 0.0, 100.0, TARIF_AWAL['Mati Total'], 1.0, key='t_mat')
    with c4:
        tarif_input['Promo'] = st.number_input(
            "Promo (%)", 0.0, 100.0, TARIF_AWAL['Promo'], 1.0, key='t_pro')

    c5, c6, c7 = st.columns([1, 1, 1.6])
    with c5:
        tarif_lain = st.number_input(
            "Tanpa kata kunci (%)", 0.0, 100.0, TARIF_DEFAULT_AWAL, 1.0, key='t_lain')
    with c6:
        tarif_flat = st.number_input(
            "Tarif pembanding (%)", 0.0, 100.0, TARIF_PEMBANDING_AWAL, 1.0, key='t_flat')
    with c7:
        prioritas = st.selectbox(
            "Kalau satu nama mengandung 2 kata kunci, yang menang:",
            ['Normal', 'Promo', 'Mati Total', 'Interface'], index=0, key='t_prio')

    st.divider()
    k1, k2 = st.columns(2)
    with k1:
        teks_kecuali = st.text_area(
            "Jasa yang tidak dihitung (satu pola per baris)",
            value="\n".join(POLA_JASA_DIKECUALIKAN), height=90, key='teks_kecuali')
    with k2:
        cab_kerusakan = st.multiselect(
            "Cabang yang kualifikasinya dibaca dari KERUSAKAN UTAMA",
            options=sorted(jasa_all['CABANG'].dropna().unique().tolist()),
            default=[c for c in CABANG_ACUAN_KERUSAKAN
                     if c in set(jasa_all['CABANG'].dropna())],
            key='cab_kerusakan')

    st.divider()
    if 'tabel_khusus' not in st.session_state:
        st.session_state['tabel_khusus'] = tarif_khusus_awal()
    tabel_khusus = st.data_editor(
        st.session_state['tabel_khusus'], key='ed_khusus', num_rows='dynamic',
        use_container_width=True, hide_index=True,
        column_config={
            'Nama Teknisi': st.column_config.TextColumn(width='medium'),
            **{lbl: st.column_config.NumberColumn(f'{lbl} (%)', min_value=0.0,
                                                 max_value=100.0, step=0.5,
                                                 format='%.1f')
               for lbl in KATEGORI_TARIF}})

    if st.button("↩️ Kembalikan ke tarif awal", key='t_reset'):
        for k, v in [('t_int', 'Interface'), ('t_nor', 'Normal'),
                     ('t_mat', 'Mati Total'), ('t_pro', 'Promo')]:
            st.session_state[k] = TARIF_AWAL[v]
        st.session_state['t_lain'] = TARIF_DEFAULT_AWAL
        st.session_state['t_flat'] = TARIF_PEMBANDING_AWAL
        st.session_state['t_prio'] = 'Normal'
        st.session_state['tabel_khusus'] = tarif_khusus_awal()
        st.session_state.pop('ed_khusus', None)
        st.rerun()

urutan = [prioritas.upper()] + [k for k in KATA_KUNCI_TARIF if k != prioritas.upper()]
peta_tarif = {k: v / 100.0 for k, v in tarif_input.items()}
peta_tarif[LABEL_LAINNYA] = tarif_lain / 100.0

jasa_all = jasa_all.copy()

pola_kecuali = [p.strip().upper() for p in teks_kecuali.splitlines() if p.strip()]
if pola_kecuali:
    b = jasa_all['BARANG'].astype(str).str.upper()
    buang = pd.Series(False, index=jasa_all.index)
    for p in pola_kecuali:
        buang |= b.str.contains(re.escape(p), regex=True, na=False)
    jasa_all = jasa_all[~buang].copy()

jasa_all['TARIF_LABEL'] = jasa_all['KW_MATCH'].map(lambda s: pilih_label_tarif(s, urutan))
if cab_kerusakan and 'KERUSAKAN' in jasa_all.columns:
    m_ker = jasa_all['CABANG'].isin(cab_kerusakan)
    if m_ker.any():
        jasa_all.loc[m_ker, 'TARIF_LABEL'] = [
            label_dari_kerusakan(k, p) for k, p in
            zip(jasa_all.loc[m_ker, 'KERUSAKAN'], jasa_all.loc[m_ker, 'KAT_JUAL'])]
jasa_all['TARIF'] = jasa_all['TARIF_LABEL'].map(peta_tarif).fillna(0.0)

khusus = peta_tarif_khusus(tabel_khusus)
peta_nama = cocokkan_teknisi(jasa_all['TEKNISI'].unique(), khusus.keys())
jasa_all['TARIF_KHUSUS'] = jasa_all['TEKNISI'].map(peta_nama)
for kunci, tar in khusus.items():
    for lbl, frac in tar.items():
        m = (jasa_all['TARIF_KHUSUS'] == kunci) & (jasa_all['TARIF_LABEL'] == lbl)
        jasa_all.loc[m, 'TARIF'] = frac

jasa_all['BAGI_HASIL'] = jasa_all['TOTAL HARGA'] * jasa_all['TARIF']
jasa_all['FLAT'] = jasa_all['TOTAL HARGA'] * (tarif_flat / 100.0)

fa, fb, fc = st.columns([2.2, 1.4, 1])
periode_list = daftar_periode_gaji(jasa_all['TGL'].min(), jasa_all['TGL'].max())
opsi = ['Semua Periode'] + periode_list
with fa:
    pilih = st.selectbox(
        "Periode penggajian (cutoff tanggal 24 s/d 23)", opsi,
        index=len(opsi) - 1 if len(opsi) > 1 else 0,
        format_func=lambda x: ("Semua Periode (tanpa cutoff)" if isinstance(x, str)
                               else label_periode(x[1], x[0])),
        key='f_periode')
with fb:
    cab_opts = ['Semua Cabang'] + sorted(jasa_all['CABANG'].dropna().unique().tolist())
    f_cabang = st.selectbox("Cabang", cab_opts, key='f_cabang')
with fc:
    sembunyikan = st.checkbox("Sembunyikan baris tanpa nama teknisi", value=False,
                              key='f_hide')

jasa = jasa_all
if f_cabang != 'Semua Cabang':
    jasa = jasa[jasa['CABANG'] == f_cabang]
if isinstance(pilih, str):
    periode_txt = "Seluruh periode data (tanpa cutoff)"
    tag_file = "semua-periode"
else:
    a, b = periode_gaji(pilih[1], pilih[0])
    jasa = jasa[(jasa['TGL'] >= a) & (jasa['TGL'] <= b)]
    periode_txt = (f"{a.day} {BULAN_NAMES[a.month]} {a.year} – "
                   f"{b.day} {BULAN_NAMES[b.month]} {b.year}")
    tag_file = f"gaji-{pilih[0]}-{pilih[1]:02d}"

jasa_tampil = jasa[jasa['TEKNISI'] != 'TIDAK ADA TEKNISI'] if sembunyikan else jasa

if jasa.empty:
    st.warning("Tidak ada transaksi jasa pada periode/cabang tersebut.")
    st.stop()

omzet = jasa['TOTAL HARGA'].sum()
bh = jasa['BAGI_HASIL'].sum()
fl = jasa['FLAT'].sum()
selisih = bh - fl
n_tek = jasa.loc[jasa['TEKNISI'] != 'TIDAK ADA TEKNISI', 'TEKNISI'].nunique()
tanpa_nama = jasa.loc[jasa['TEKNISI'] == 'TIDAK ADA TEKNISI', 'TOTAL HARGA'].sum()

st.markdown(kpi_html([
    {'label': 'Omzet Jasa', 'value': rp(omzet), 'sub': f"{len(jasa):,} baris",
     'grad': 'linear-gradient(135deg,#1f3864,#2e5394)'},
    {'label': 'Bagi Hasil (Aturan)', 'value': rp(bh),
     'sub': f"{(bh/omzet*100 if omzet else 0):.1f}% dari omzet jasa",
     'grad': 'linear-gradient(135deg,#16a34a,#22c55e)'},
    {'label': f'Pembanding Flat {tarif_flat:.0f}%', 'value': rp(fl),
     'sub': f'omzet jasa × {tarif_flat:.0f}%',
     'grad': 'linear-gradient(135deg,#7c3aed,#a855f7)'},
    {'label': 'Selisih', 'value': rp(selisih),
     'sub': ('aturan lebih besar' if selisih > 0
             else 'flat lebih besar' if selisih < 0 else 'sama'),
     'grad': ('linear-gradient(135deg,#e0921f,#e2b21a)' if selisih >= 0
              else 'linear-gradient(135deg,#c9392f,#e0475a)')},
    {'label': 'Jumlah Teknisi', 'value': f"{n_tek:,}",
     'sub': f"rata-rata {rp(bh/n_tek if n_tek else 0)}/teknisi",
     'grad': 'linear-gradient(135deg,#0f8a82,#17a3a3)'},
    {'label': 'Omzet Tanpa Nama Teknisi', 'value': rp(tanpa_nama),
     'sub': f"{(jasa['TEKNISI'] == 'TIDAK ADA TEKNISI').sum():,} baris",
     'grad': 'linear-gradient(135deg,#64748b,#94a3b8)'},
]), unsafe_allow_html=True)
st.write("")

lbl_flat = f'Pembanding {tarif_flat:.0f}%'

st.markdown("### Rekap Bagi Hasil per Teknisi & Cabang")
rek = (jasa_tampil.groupby(['TEKNISI', 'CABANG'], as_index=False)
       .agg(Baris=('TOTAL HARGA', 'size'),
            Omzet_Jasa=('TOTAL HARGA', 'sum'),
            Bagi_Hasil=('BAGI_HASIL', 'sum'),
            Flat=('FLAT', 'sum')))
rek['Selisih'] = rek['Bagi_Hasil'] - rek['Flat']
rek['Efektif %'] = (rek['Bagi_Hasil'] / rek['Omzet_Jasa'] * 100).round(1)
rek = rek.sort_values('Bagi_Hasil', ascending=False)

rek_show = rek.rename(columns={
    'TEKNISI': 'Nama Teknisi', 'CABANG': 'Cabang',
    'Omzet_Jasa': 'Omzet Jasa', 'Bagi_Hasil': 'Bagi Hasil (Aturan)', 'Flat': lbl_flat})

cari = st.text_input("Cari nama teknisi / cabang", key='cari_rekap')
rek_view = rek_show
if cari:
    m = rek_show.apply(lambda r: cari.upper() in
                       f"{r['Nama Teknisi']} {r['Cabang']}".upper(), axis=1)
    rek_view = rek_show[m]

st.dataframe(
    rek_view.style.format({
        'Baris': '{:,.0f}', 'Omzet Jasa': 'Rp {:,.0f}',
        'Bagi Hasil (Aturan)': 'Rp {:,.0f}', lbl_flat: 'Rp {:,.0f}',
        'Selisih': 'Rp {:,.0f}'}),
    use_container_width=True, height=460, hide_index=True, key='tabel_rekap')

KATEGORI_ORDER = ['Interface', 'Normal', 'Mati Total', 'Promo', LABEL_LAINNYA]

KOLOM_POTONGAN = ['Potongan Refund', 'Potongan AR', 'Potongan Kasbon', 'Keterlambatan',
                  'Potongan Minus Audit', 'Potongan Audit Compliance',
                  'Biaya Pendaftaran Koperasi', 'Simpanan Pokok', 'Simpanan Wajib']
KOLOM_CADANGAN = ['Cadangan 7 Tahun / bulan', 'Cadangan 7 Tahun']
KOLOM_RUMUS = ['Total Potongan', 'Gaji Teknisi', 'Nett Bagi hasil',
               'Total Cadangan 7 Tahun']
KOLOM_GAJI = (KOLOM_POTONGAN + ['Total Potongan', 'Gaji Teknisi', 'Nett Bagi hasil']
              + KOLOM_CADANGAN + ['Total Cadangan 7 Tahun'])


def _sheet_name(nama, terpakai):
    s = str(nama)
    for ch in '[]:*?/\\':
        s = s.replace(ch, '-')
    s = s.strip() or 'Cabang'
    s = s[:31]
    dasar, n = s, 2
    while s.lower() in terpakai:
        akhiran = f"_{n}"
        s = dasar[:31 - len(akhiran)] + akhiran
        n += 1
    terpakai.add(s.lower())
    return s


def rekap_kualifikasi(df, keys):
    base = (df.groupby(keys, as_index=False)
              .agg(Baris=('TOTAL HARGA', 'size'),
                   Omzet=('TOTAL HARGA', 'sum'),
                   BH=('BAGI_HASIL', 'sum'),
                   Flat=('FLAT', 'sum')))

    def _pivot(nilai, prefix):
        p = df.pivot_table(index=keys, columns='TARIF_LABEL', values=nilai,
                           aggfunc='sum', fill_value=0.0)
        for k in KATEGORI_ORDER:
            if k not in p.columns:
                p[k] = 0.0
        p = p[KATEGORI_ORDER]
        p.columns = [f"{prefix} {k}" for k in KATEGORI_ORDER]
        return p.reset_index()

    out = (base.merge(_pivot('TOTAL HARGA', 'Omzet'), on=keys, how='left')
               .merge(_pivot('BAGI_HASIL', 'Bagi Hasil'), on=keys, how='left'))
    out['Selisih'] = out['BH'] - out['Flat']
    out['Efektif %'] = (out['BH'] / out['Omzet'].replace(0, pd.NA) * 100).round(1)

    urut = list(keys) + ['Baris'] \
        + [f"Omzet {k}" for k in KATEGORI_ORDER] + ['Omzet'] \
        + [f"Bagi Hasil {k}" for k in KATEGORI_ORDER] + ['BH', 'Flat', 'Selisih', 'Efektif %']
    out = out[urut].rename(columns={
        'TEKNISI': 'Nama Teknisi', 'CABANG': 'Cabang',
        'Omzet': 'Omzet Jasa (Total)', 'BH': 'Bagi Hasil (Aturan)', 'Flat': lbl_flat})
    return out.sort_values('Bagi Hasil (Aturan)', ascending=False).reset_index(drop=True)


def _rumus_gaji(df, r):
    def kol(nama):
        return get_column_letter(df.columns.get_loc(nama) + 1)

    return {
        'Total Potongan':
            f"=SUM({kol(KOLOM_POTONGAN[0])}{r}:{kol(KOLOM_POTONGAN[-1])}{r})",
        'Nett Bagi hasil':
            f"={kol('Bagi Hasil (Aturan)')}{r}-{kol('Total Potongan')}{r}",
        'Total Cadangan 7 Tahun':
            f"={kol('Cadangan 7 Tahun / bulan')}{r}+{kol('Cadangan 7 Tahun')}{r}",
    }


def salin_sheet(ws_src, wb_dst, title_dst=None):
    """Menyalin sheet dari source workbook ke destination workbook beserta gaya sel & lebar kolom."""
    if title_dst is None:
        title_dst = ws_src.title
    ws_dst = wb_dst.create_sheet(title=title_dst)

    # 1. Salin data & styling sel
    for row in ws_src.iter_rows():
        for cell in row:
            val = norm_formula(cell.value)
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=val)
            if cell.has_style:
                try: new_cell.font = copy(cell.font)
                except Exception: pass
                try: new_cell.border = copy(cell.border)
                except Exception: pass
                try: new_cell.fill = copy(cell.fill)
                except Exception: pass
                if cell.number_format:
                    new_cell.number_format = cell.number_format
                try: new_cell.protection = copy(cell.protection)
                except Exception: pass
                try: new_cell.alignment = copy(cell.alignment)
                except Exception: pass

    # 2. Salin lebar kolom bawaan
    for col_key, col_dim in ws_src.column_dimensions.items():
        if col_dim.width is None:
            continue
        col_str = str(col_key)
        if ':' in col_str:
            parts = col_str.split(':')
            try:
                start_idx = int(parts[0]) if parts[0].isdigit() else column_index_from_string(parts[0])
                end_idx = int(parts[1]) if parts[1].isdigit() else column_index_from_string(parts[1])
                for col_idx in range(start_idx, end_idx + 1):
                    ws_dst.column_dimensions[get_column_letter(col_idx)].width = col_dim.width
            except Exception:
                pass
        else:
            try:
                letter = get_column_letter(int(col_str)) if col_str.isdigit() else col_str
                ws_dst.column_dimensions[letter].width = col_dim.width
            except Exception:
                pass

    # 3. Salin tinggi baris
    for row_idx, row_dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row_idx].height = row_dim.height

    # 4. Salin merged cells
    for merged_cell in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged_cell))

    # 5. Salin Pengaturan Halaman & Margin Cetak
    if hasattr(ws_src, 'page_setup') and ws_src.page_setup:
        ws_dst.page_setup.orientation = ws_src.page_setup.orientation
        ws_dst.page_setup.paperSize = ws_src.page_setup.paperSize
        ws_dst.page_setup.fitToWidth = 1
        ws_dst.page_setup.fitToHeight = 1
        ws_dst.sheet_properties.pageSetUpPr.fitToPage = True

    if hasattr(ws_src, 'page_margins') and ws_src.page_margins:
        ws_dst.page_margins.left = ws_src.page_margins.left
        ws_dst.page_margins.right = ws_src.page_margins.right
        ws_dst.page_margins.top = ws_src.page_margins.top
        ws_dst.page_margins.bottom = ws_src.page_margins.bottom
        ws_dst.page_margins.header = ws_src.page_margins.header
        ws_dst.page_margins.footer = ws_src.page_margins.footer

    # Khusus sheet 'FINAL' / 'final': atur fixed lebar kolom P s/d Y = 42.57
    if str(title_dst).strip().lower() == 'final':
        for col_letter in ['P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']:
            ws_dst.column_dimensions[col_letter].width = 42.57

    return ws_dst


def _tulis_sheet(wb, df, nama_sheet, judul, kolom_gaji=False):
    ws = wb.create_sheet(title=nama_sheet)

    ws.cell(row=1, column=1, value=judul).font = Font(bold=True, size=13, color='1F3864')
    ws.cell(row=2, column=1,
            value=f"Periode: {periode_txt} · Tarif: "
                  + ", ".join(f"{k} {v:.0f}%" for k, v in tarif_input.items())
                  + f", Lainnya {tarif_lain:.0f}%, pembanding flat {tarif_flat:.0f}%"
            ).font = Font(size=9, italic=True, color='555555')

    n_baris, n_kol = len(df), len(df.columns)
    head_fill_biru = PatternFill('solid', fgColor='1F3864')
    head_fill_coklat = PatternFill('solid', fgColor='B45309')
    head_fill_kuning = PatternFill('solid', fgColor='EAB308')
    head_fill_hijau = PatternFill('solid', fgColor='166534')

    head_font_putih = Font(bold=True, color='FFFFFF', size=10)
    head_font_hitam = Font(bold=True, color='000000', size=10)
    thin = Side(style='thin', color='D9D9D9')

    for j, kol in enumerate(df.columns, start=1):
        c = ws.cell(row=4, column=j, value=kol)
        c.font = head_font_putih

        if kol.startswith('Bagi Hasil'):
            c.fill = head_fill_kuning
            c.font = head_font_hitam
        elif kol in (KOLOM_POTONGAN + KOLOM_CADANGAN + ['Total Potongan']):
            c.fill = head_fill_coklat
        elif kol in (KOLOM_RUMUS + ['Gaji Teknisi']):
            c.fill = head_fill_hijau
        else:
            c.fill = head_fill_biru

        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(bottom=Side(style='medium', color='1F3864'))
        lebar = max(len(str(kol)) + 2,
                    (df[kol].astype(str).str.len().max() if n_baris else 0) + 2)
        ws.column_dimensions[get_column_letter(j)].width = min(max(lebar, 10), 34)

    kol_rp = [c for c in df.columns
              if c.startswith(('Omzet', 'Bagi Hasil')) or c in (lbl_flat, 'Selisih')
              or c in KOLOM_GAJI]
    idx_rp = [df.columns.get_loc(c) + 1 for c in kol_rp]
    idx_baris = (df.columns.get_loc('Baris') + 1) if 'Baris' in df.columns else None
    idx_pct = (df.columns.get_loc('Efektif %') + 1) if 'Efektif %' in df.columns else None

    def kol_letter(nama_kolom):
        return get_column_letter(df.columns.get_loc(nama_kolom) + 1)

    for i in range(n_baris):
        r = 5 + i
        for j, kol in enumerate(df.columns, start=1):
            val = df.iloc[i][kol]
            ws.cell(row=r, column=j, value=val if pd.notna(val) else None)

        tek_nama = df.iloc[i].get('Nama Teknisi', '')
        kunci_tek = peta_nama.get(tek_nama)
        for k in KATEGORI_ORDER:
            omzet_k = f"Omzet {k}"
            bh_k = f"Bagi Hasil {k}"
            if omzet_k in df.columns and bh_k in df.columns:
                tar_frac = peta_tarif.get(k, 0.3)
                if kunci_tek and kunci_tek in khusus and k in khusus[kunci_tek]:
                    tar_frac = khusus[kunci_tek][k]
                ws.cell(row=r, column=df.columns.get_loc(bh_k) + 1,
                        value=f"={kol_letter(omzet_k)}{r}*{tar_frac}")

        if 'Omzet Jasa (Total)' in df.columns and 'Omzet Interface' in df.columns:
            ws.cell(row=r, column=df.columns.get_loc('Omzet Jasa (Total)') + 1,
                    value=f"=SUM({kol_letter('Omzet Interface')}{r}:{kol_letter('Omzet Lainnya')}{r})")

        if 'Bagi Hasil (Aturan)' in df.columns and 'Bagi Hasil Interface' in df.columns:
            ws.cell(row=r, column=df.columns.get_loc('Bagi Hasil (Aturan)') + 1,
                    value=f"=SUM({kol_letter('Bagi Hasil Interface')}{r}:{kol_letter('Bagi Hasil Lainnya')}{r})")

        if lbl_flat in df.columns and 'Omzet Jasa (Total)' in df.columns:
            ws.cell(row=r, column=df.columns.get_loc(lbl_flat) + 1,
                    value=f"={kol_letter('Omzet Jasa (Total)')}{r}*{tarif_flat/100}")

        if 'Selisih' in df.columns and 'Bagi Hasil (Aturan)' in df.columns and lbl_flat in df.columns:
            ws.cell(row=r, column=df.columns.get_loc('Selisih') + 1,
                    value=f"={kol_letter('Bagi Hasil (Aturan)')}{r}-{kol_letter(lbl_flat)}{r}")

        if 'Efektif %' in df.columns and 'Omzet Jasa (Total)' in df.columns and 'Bagi Hasil (Aturan)' in df.columns:
            ws.cell(row=r, column=df.columns.get_loc('Efektif %') + 1,
                    value=f"=IF({kol_letter('Omzet Jasa (Total)')}{r}=0,0,{kol_letter('Bagi Hasil (Aturan)')}{r}/{kol_letter('Omzet Jasa (Total)')}{r}*100)")

        for j in idx_rp:
            ws.cell(row=r, column=j).number_format = '#,##0'
        if idx_baris:
            ws.cell(row=r, column=idx_baris).number_format = '#,##0'
        if idx_pct:
            ws.cell(row=r, column=idx_pct).number_format = '0.0'
        for j in range(1, n_kol + 1):
            ws.cell(row=r, column=j).border = Border(bottom=thin)

        if kolom_gaji:
            for kol in KOLOM_POTONGAN + KOLOM_CADANGAN + ['Total Potongan']:
                ws.cell(row=r, column=df.columns.get_loc(kol) + 1).fill = \
                    PatternFill('solid', fgColor='FFF8E1')
            for kol, rumus in _rumus_gaji(df, r).items():
                sel = ws.cell(row=r, column=df.columns.get_loc(kol) + 1, value=norm_formula(rumus))
                sel.number_format = '#,##0'

    if n_baris:
        rt = 5 + n_baris
        ws.cell(row=rt, column=1, value='TOTAL')
        for j in range(1, n_kol + 1):
            c = ws.cell(row=rt, column=j)
            c.font = Font(bold=True)
            c.fill = PatternFill('solid', fgColor='DCE6F1')
            c.border = Border(top=Side(style='medium', color='1F3864'))
        for j in idx_rp + ([idx_baris] if idx_baris else []):
            L = get_column_letter(j)
            c = ws.cell(row=rt, column=j, value=f"=SUM({L}5:{L}{rt-1})")
            c.number_format = '#,##0'
            c.font = Font(bold=True)
        if idx_pct and 'Bagi Hasil (Aturan)' in df.columns:
            Lb = get_column_letter(df.columns.get_loc('Bagi Hasil (Aturan)') + 1)
            Lo = get_column_letter(df.columns.get_loc('Omzet Jasa (Total)') + 1)
            c = ws.cell(row=rt, column=idx_pct,
                        value=f"=IF({Lo}{rt}=0,0,{Lb}{rt}/{Lo}{rt}*100)")
            c.number_format = '0.0'
            c.font = Font(bold=True)

    ws.freeze_panes = ws.cell(row=5, column=2)
    if n_baris:
        ws.auto_filter.ref = f"A4:{get_column_letter(n_kol)}{4 + n_baris}"


def buat_excel(df_sumber, raw_bytes=None, nama_cabang_file='Semua Cabang'):
    """Workbook: Sheet 1 (Rincian Faktur Penjualan), Sheet 2 (Pivot), Sheet 3 (RAW), Sheet 4 (FINAL/final)."""
    buf = io.BytesIO()

    d = df_sumber.copy()
    d['TEKNISI'] = (d['TEKNISI'].fillna('TIDAK ADA TEKNISI').astype(str).str.strip()
                    .replace({'': 'TIDAK ADA TEKNISI', 'nan': 'TIDAK ADA TEKNISI',
                              'NaN': 'TIDAK ADA TEKNISI', 'None': 'TIDAK ADA TEKNISI'}))
    d['CABANG'] = (d['CABANG'].fillna('(TANPA CABANG)').astype(str).str.strip()
                   .replace({'': '(TANPA CABANG)', 'nan': '(TANPA CABANG)'}))

    if TEMPLATE_PATH.exists():
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        if raw_bytes and 'Rincian Faktur Penjualan' in wb.sheetnames:
            ws_rincian = wb['Rincian Faktur Penjualan']
            wb_up = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
            ws_up = wb_up.active

            if ws_rincian.max_row > 1:
                ws_rincian.delete_rows(2, ws_rincian.max_row)

            for r in range(2, ws_up.max_row + 1):
                row_vals = [ws_up.cell(r, c).value for c in range(1, 47)]
                if any(v is not None for v in row_vals):
                    formula_katakunci = f'=_xlfn.IFS(ISNUMBER(SEARCH("Interface", AJ{r})), "Omset Interface", ISNUMBER(SEARCH("Normal", AJ{r})), "Omset Normal", ISNUMBER(SEARCH("Mati Total", AJ{r})), "Omset Mati Total", ISNUMBER(SEARCH("Promo", AJ{r})), "Omset Promo", TRUE, "Omset lainnya")'
                    row_vals.append(norm_formula(formula_katakunci))

                    ws_rincian.append(row_vals)
    else:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Rincian Faktur Penjualan"
        wb.create_sheet(title="Pivot")

    terpakai = set(s.lower() for s in wb.sheetnames)
    cabang_list = sorted(d['CABANG'].unique())

    n_data_raw = 0
    for idx, cab in enumerate(cabang_list):
        sub = d[d['CABANG'] == cab]
        if sub.empty:
            continue
        dc = rekap_kualifikasi(sub, ['TEKNISI']).copy()
        dc.insert(1, 'Cabang', cab)
        for kol in KOLOM_GAJI:
            dc[kol] = pd.NA

        # Sheet ketiga selalu dinamai 'RAW'
        if idx == 0:
            s_name = 'RAW'
            terpakai.add('raw')
            n_data_raw = len(dc)
        else:
            s_name = _sheet_name(cab, terpakai)

        _tulis_sheet(wb, dc, s_name, f'Bagi Hasil Teknisi — Cabang {cab}', kolom_gaji=True)

    # Tambahkan Sheet ke-4 dari master file (menyesuaikan nama sheet di master file)
    if FINAL_PATH and FINAL_PATH.exists():
        try:
            wb_final = openpyxl.load_workbook(FINAL_PATH)
            sheet_target = None
            for s in wb_final.sheetnames:
                if s.strip().lower() == 'final':
                    sheet_target = wb_final[s]
                    break
            if sheet_target is None:
                sheet_target = wb_final.active

            # 1. Salin sheet (nama sheet mengikuti master file)
            ws_f = salin_sheet(sheet_target, wb, title_dst=sheet_target.title)

            # 2. Update Nama Cabang di Baris 2 (Sel B2)
            val_b2 = str(ws_f.cell(row=2, column=2).value or '')
            if val_b2 and nama_cabang_file and nama_cabang_file != 'Semua Cabang':
                new_b2 = re.sub(r'MFLASH\s+[A-Z0-9\_]+', f'MFLASH {nama_cabang_file}', val_b2, flags=re.I)
                if new_b2 == val_b2:
                    new_b2 = val_b2.replace('JATIMULYA', nama_cabang_file)
                ws_f.cell(row=2, column=2, value=new_b2)

            # 3. Update Periode Bulan & Tahun di Baris 4 (Sel D4)
            if 'TGL' in df_sumber.columns and not df_sumber['TGL'].dropna().empty:
                t_max_f = df_sumber['TGL'].max()
                if pd.notna(t_max_f):
                    bln_nama = BULAN_NAMES[t_max_f.month].upper()
                    thn_num = t_max_f.year
                    ws_f.cell(row=4, column=4, value=f"{bln_nama} {thn_num}")

            # 4. Sinkronisasi Baris Data dengan Sheet RAW
            baris_header = 7
            baris_awal_data = 8
            col_no = 2
            col_nama = 4

            # Cari posisi baris TOTAL di sheet final
            baris_total = None
            for r in range(baris_awal_data, ws_f.max_row + 1):
                val = str(ws_f.cell(r, col_no).value or ws_f.cell(r, col_nama).value or '').strip().upper()
                if 'TOTAL' in val:
                    baris_total = r
                    break

            if baris_total and n_data_raw > 0:
                n_template_rows = baris_total - baris_awal_data

                # A. Jika data RAW lebih sedikit dari template: hapus baris berlebih
                if n_template_rows > n_data_raw:
                    baris_hapus_mulai = baris_awal_data + n_data_raw
                    jumlah_dihapus = n_template_rows - n_data_raw
                    ws_f.delete_rows(baris_hapus_mulai, jumlah_dihapus)

                # B. Jika data RAW lebih banyak dari template: sisipkan baris baru
                elif n_data_raw > n_template_rows:
                    jumlah_disisip = n_data_raw - n_template_rows
                    ws_f.insert_rows(baris_total, amount=jumlah_disisip)

                # Update & salin nilai serta rumus untuk seluruh baris data
                for idx_s in range(n_data_raw):
                    r_curr = baris_awal_data + idx_s
                    r_raw = 5 + idx_s

                    for col_idx in range(1, ws_f.max_column + 1):
                        ref_cell = ws_f.cell(baris_awal_data, col_idx)
                        new_cell = ws_f.cell(r_curr, col_idx)

                        if ref_cell.has_style:
                            new_cell.font = copy(ref_cell.font)
                            new_cell.border = copy(ref_cell.border)
                            new_cell.fill = copy(ref_cell.fill)
                            new_cell.number_format = ref_cell.number_format
                            new_cell.alignment = copy(ref_cell.alignment)

                        if col_idx == col_no:
                            new_cell.value = idx_s + 1
                        elif ref_cell.value and isinstance(ref_cell.value, str) and 'RAW!' in ref_cell.value:
                            val = re.sub(r'(RAW![A-Z]+)\d+', rf'\g<1>{r_raw}', ref_cell.value)
                            new_cell.value = norm_formula(val)
                        elif ref_cell.value and isinstance(ref_cell.value, str) and ref_cell.value.startswith('='):
                            val = re.sub(rf'([A-Z]+){baris_awal_data}\b', rf'\g<1>{r_curr}', ref_cell.value)
                            new_cell.value = norm_formula(val)
                        elif ref_cell.value is not None:
                            new_cell.value = ref_cell.value

                # Re-number kolom NO untuk seluruh baris data
                for idx_n in range(n_data_raw):
                    r_curr = baris_awal_data + idx_n
                    ws_f.cell(row=r_curr, column=col_no, value=idx_n + 1)

                # 5. Perbarui baris TOTAL dan rumus SUM
                baris_total_baru = baris_awal_data + n_data_raw
                baris_data_akhir = baris_total_baru - 1

                for col_idx in range(1, ws_f.max_column + 1):
                    cell = ws_f.cell(row=baris_total_baru, column=col_idx)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        val = re.sub(
                            r'SUM\(([A-Z]+)\d+:([A-Z]+)\d+\)',
                            rf'SUM(\g<1>{baris_awal_data}:\g<2>{baris_data_akhir})',
                            cell.value,
                            flags=re.I
                        )
                        cell.value = norm_formula(val)

        except Exception as e:
            st.warning(f"Gagal menyalin/menyesuaikan sheet dari master file: {e}")

    # 6. Format / Beri Warna pada Sheet 'Pivot'
    if 'Pivot' in wb.sheetnames:
        ws_pvt = wb['Pivot']
        head_fill_biru = PatternFill('solid', fgColor='1F3864')
        head_font_putih = Font(bold=True, color='FFFFFF', size=10)
        total_fill_biru = PatternFill('solid', fgColor='DCE6F1')
        total_font_biru = Font(bold=True, color='1F3864', size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        max_c = max(ws_pvt.max_column, 6)
        for r in (3, 4):
            for c in range(1, max_c + 1):
                cell = ws_pvt.cell(r, c)
                cell.fill = head_fill_biru
                cell.font = head_font_putih
                cell.alignment = Alignment(horizontal='center' if c > 1 else 'left', vertical='center')

        for r in range(5, max(ws_pvt.max_row + 1, 30)):
            val_first = str(ws_pvt.cell(r, 1).value or '').strip().upper()
            is_grand_total = 'GRAND TOTAL' in val_first or 'TOTAL' in val_first
            for c in range(1, max_c + 1):
                cell = ws_pvt.cell(r, c)
                if is_grand_total:
                    cell.fill = total_fill_biru
                    cell.font = total_font_biru
                if c > 1 and cell.value is not None:
                    cell.number_format = '#,##0'
                if cell.value is not None or is_grand_total:
                    cell.border = thin_border

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# Menentukan nama file download otomatis
if f_cabang != 'Semua Cabang':
    nama_cabang_file = f_cabang
else:
    cabs_file = sorted(jasa_tampil['CABANG'].dropna().unique().tolist())
    nama_cabang_file = cabs_file[0] if len(cabs_file) == 1 else 'Semua Cabang'

if isinstance(pilih, tuple):
    a_f, b_f = periode_gaji(pilih[1], pilih[0])
    label_tgl_file = f"{a_f.day} {BULAN_NAMES[a_f.month]} - {b_f.day} {BULAN_NAMES[b_f.month]} {b_f.year}"
else:
    if not jasa_tampil.empty and 'TGL' in jasa_tampil.columns:
        t_min_f = jasa_tampil['TGL'].min()
        t_max_f = jasa_tampil['TGL'].max()
        if pd.notna(t_min_f) and pd.notna(t_max_f):
            label_tgl_file = f"{t_min_f.day} {BULAN_NAMES[t_min_f.month]} - {t_max_f.day} {BULAN_NAMES[t_max_f.month]} {t_max_f.year}"
        else:
            label_tgl_file = date.today().strftime('%d-%m-%Y')
    else:
        label_tgl_file = date.today().strftime('%d-%m-%Y')

nama_file_download = f"Bagi hasil teknisi {label_tgl_file} {nama_cabang_file}.xlsx"

with st.container():
    st.markdown("##### 📊 Unduh Excel (Sesuai Template + Sheet RAW + Sheet FINAL)")
    st.caption(
        "Mengunduh file Excel yang berisi **Sheet 1 (Rincian Faktur Penjualan)**, "
        "**Sheet 2 (Pivot)**, **Sheet 3 (RAW)**, dan **Sheet 4 (FINAL)**."
    )
    if st.button("🧾 Siapkan berkas Excel", key='siap_xlsx', use_container_width=True):
        with st.spinner("Menyusun workbook..."):
            st.session_state['xlsx_bytes'] = buat_excel(jasa_tampil, raw_uploaded_bytes, nama_cabang_file)
            st.session_state['xlsx_filename'] = nama_file_download
    if st.session_state.get('xlsx_bytes') is not None:
        st.download_button(
            "⬇️ Unduh Excel (.xlsx)",
            data=st.session_state['xlsx_bytes'],
            file_name=st.session_state.get('xlsx_filename', nama_file_download),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key='unduh_xlsx')